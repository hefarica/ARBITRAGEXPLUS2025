"""
ExcelClient - Cliente para lectura/escritura de Excel con detección automática de columnas PUSH/PULL
basada en colores de encabezados.

Convención de colores:
- AZUL (RGB: 0, 112, 192 o similar) = PUSH (sistema escribe, usuario lee)
- BLANCO (RGB: 255, 255, 255 o sin color) = PULL (usuario escribe, sistema lee)
"""

import logging
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ColumnMode(Enum):
    """Modo de operación de una columna"""
    PUSH = "PUSH"  # Sistema escribe, usuario lee
    PULL = "PULL"  # Usuario escribe, sistema lee
    UNKNOWN = "UNKNOWN"  # No se pudo determinar


@dataclass
class ColumnMetadata:
    """Metadata de una columna en Excel"""
    index: int  # Índice de columna (0-based)
    letter: str  # Letra de columna (A, B, C, ...)
    name: str  # Nombre del encabezado
    mode: ColumnMode  # Modo PUSH/PULL
    bg_color: Optional[str] = None  # Color de fondo en formato RGB hex


class ExcelClient:
    """Cliente para operaciones con archivos Excel con detección automática de columnas PUSH/PULL"""

    # Colores de referencia (RGB)
    BLUE_PUSH_COLORS = [
        "FF0070C0",  # Azul estándar de Excel (0, 112, 192)
        "000070C0",  # Variante sin alpha
        "0070C0",    # Variante corta
        "4472C4",    # Azul tema de Office
        "5B9BD5",    # Azul claro tema
    ]

    WHITE_PULL_COLORS = [
        "FFFFFFFF",  # Blanco con alpha
        "00FFFFFF",  # Blanco sin alpha
        "FFFFFF",    # Blanco corto
        None,        # Sin color (por defecto)
    ]

    def __init__(self, file_path: str | Path):
        """
        Inicializa el cliente Excel.

        Args:
            file_path: Ruta al archivo Excel
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self._column_cache: Dict[str, List[ColumnMetadata]] = {}

        if not self.file_path.exists():
            raise FileNotFoundError(f"Archivo Excel no encontrado: {self.file_path}")

        logger.info(f"ExcelClient inicializado para: {self.file_path}")

    def load(self) -> None:
        """Carga el archivo Excel en memoria"""
        try:
            self.workbook = load_workbook(filename=self.file_path, data_only=False)
            logger.info(f"Excel cargado: {len(self.workbook.sheetnames)} hojas")
        except Exception as e:
            logger.error(f"Error al cargar Excel: {e}")
            raise

    def close(self) -> None:
        """Cierra el workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            logger.debug("Workbook cerrado")

    def save(self, file_path: Optional[str | Path] = None) -> None:
        """
        Guarda el workbook.

        Args:
            file_path: Ruta de destino (opcional, por defecto sobrescribe el original)
        """
        if not self.workbook:
            raise ValueError("No hay workbook cargado")

        save_path = Path(file_path) if file_path else self.file_path
        self.workbook.save(save_path)
        logger.info(f"Excel guardado en: {save_path}")

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Obtiene una hoja por nombre.

        Args:
            sheet_name: Nombre de la hoja

        Returns:
            Worksheet de openpyxl
        """
        if not self.workbook:
            raise ValueError("Workbook no cargado. Llama a load() primero.")

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Hoja '{sheet_name}' no existe. Hojas disponibles: {self.workbook.sheetnames}")

        return self.workbook[sheet_name]

    def _detect_column_mode(self, cell_color: Optional[Color]) -> ColumnMode:
        """
        Detecta el modo de columna basándose en el color de fondo.

        Args:
            cell_color: Color de la celda (openpyxl Color object)

        Returns:
            ColumnMode detectado
        """
        if cell_color is None or cell_color.rgb is None:
            # Sin color = PULL (blanco por defecto)
            return ColumnMode.PULL

        rgb_value = cell_color.rgb

        # Verificar si es azul (PUSH)
        for blue in self.BLUE_PUSH_COLORS:
            if blue and rgb_value.upper() == blue.upper():
                return ColumnMode.PUSH

        # Verificar si es blanco (PULL)
        for white in self.WHITE_PULL_COLORS:
            if white and rgb_value.upper() == white.upper():
                return ColumnMode.PULL

        # Si no coincide con ninguno, intentar detectar por componentes RGB
        if rgb_value.startswith("FF") or rgb_value.startswith("00"):
            rgb_hex = rgb_value[2:]  # Quitar alpha
        else:
            rgb_hex = rgb_value

        # Convertir a RGB
        try:
            if len(rgb_hex) == 6:
                r = int(rgb_hex[0:2], 16)
                g = int(rgb_hex[2:4], 16)
                b = int(rgb_hex[4:6], 16)

                # Azul: componente B alta, R y G bajas
                if b > 150 and r < 100 and g < 150:
                    return ColumnMode.PUSH

                # Blanco: todos los componentes altos
                if r > 240 and g > 240 and b > 240:
                    return ColumnMode.PULL
        except ValueError:
            logger.warning(f"No se pudo parsear color RGB: {rgb_value}")

        logger.warning(f"Color no reconocido: {rgb_value}, asumiendo UNKNOWN")
        return ColumnMode.UNKNOWN

    def get_column_metadata(self, sheet_name: str, header_row: int = 1) -> List[ColumnMetadata]:
        """
        Obtiene metadata de todas las columnas de una hoja basándose en los encabezados.

        Args:
            sheet_name: Nombre de la hoja
            header_row: Fila de encabezados (1-based, por defecto 1)

        Returns:
            Lista de ColumnMetadata para cada columna
        """
        # Verificar cache
        cache_key = f"{sheet_name}_{header_row}"
        if cache_key in self._column_cache:
            logger.debug(f"Usando metadata en cache para {cache_key}")
            return self._column_cache[cache_key]

        sheet = self.get_sheet(sheet_name)
        metadata_list: List[ColumnMetadata] = []

        # Iterar sobre las columnas hasta encontrar una vacía
        col_idx = 0
        for col in sheet.iter_cols(min_row=header_row, max_row=header_row):
            cell = col[0]

            # Si la celda está vacía, asumimos que terminaron las columnas
            if cell.value is None or str(cell.value).strip() == "":
                break

            # Obtener color de fondo
            bg_color = cell.fill.fgColor if cell.fill else None
            mode = self._detect_column_mode(bg_color)

            metadata = ColumnMetadata(
                index=col_idx,
                letter=get_column_letter(col_idx + 1),
                name=str(cell.value).strip(),
                mode=mode,
                bg_color=bg_color.rgb if bg_color and bg_color.rgb else None
            )

            metadata_list.append(metadata)
            col_idx += 1

        logger.info(f"Detectadas {len(metadata_list)} columnas en '{sheet_name}'")
        logger.debug(f"PUSH: {sum(1 for m in metadata_list if m.mode == ColumnMode.PUSH)}, "
                    f"PULL: {sum(1 for m in metadata_list if m.mode == ColumnMode.PULL)}, "
                    f"UNKNOWN: {sum(1 for m in metadata_list if m.mode == ColumnMode.UNKNOWN)}")

        # Guardar en cache
        self._column_cache[cache_key] = metadata_list

        return metadata_list

    def get_push_columns(self, sheet_name: str, header_row: int = 1) -> List[ColumnMetadata]:
        """Obtiene solo las columnas PUSH (sistema escribe)"""
        all_cols = self.get_column_metadata(sheet_name, header_row)
        return [col for col in all_cols if col.mode == ColumnMode.PUSH]

    def get_pull_columns(self, sheet_name: str, header_row: int = 1) -> List[ColumnMetadata]:
        """Obtiene solo las columnas PULL (usuario escribe)"""
        all_cols = self.get_column_metadata(sheet_name, header_row)
        return [col for col in all_cols if col.mode == ColumnMode.PULL]

    def get_sheet_data(
        self,
        sheet_name: str,
        header_row: int = 1,
        start_row: int = 2,
        end_row: Optional[int] = None,
        columns: Optional[List[str]] = None,
        mode_filter: Optional[ColumnMode] = None
    ) -> List[Dict[str, Any]]:
        """
        Obtiene datos de una hoja como lista de diccionarios.

        Args:
            sheet_name: Nombre de la hoja
            header_row: Fila de encabezados (1-based)
            start_row: Fila inicial de datos (1-based)
            end_row: Fila final de datos (1-based, None = hasta el final)
            columns: Lista de nombres de columnas a incluir (None = todas)
            mode_filter: Filtrar solo columnas PUSH o PULL (None = todas)

        Returns:
            Lista de diccionarios {nombre_columna: valor}
        """
        sheet = self.get_sheet(sheet_name)
        metadata = self.get_column_metadata(sheet_name, header_row)

        # Filtrar columnas según criterios
        if mode_filter:
            metadata = [m for m in metadata if m.mode == mode_filter]

        if columns:
            metadata = [m for m in metadata if m.name in columns]

        if not metadata:
            logger.warning(f"No hay columnas que cumplan los criterios en '{sheet_name}'")
            return []

        # Determinar rango de filas
        max_row = sheet.max_row if end_row is None else end_row
        data_rows: List[Dict[str, Any]] = []

        for row_idx in range(start_row, max_row + 1):
            row_data = {}
            has_data = False

            for col_meta in metadata:
                cell = sheet.cell(row=row_idx, column=col_meta.index + 1)
                value = cell.value

                if value is not None:
                    has_data = True

                row_data[col_meta.name] = value

            # Solo agregar filas que tengan al menos un dato
            if has_data:
                data_rows.append(row_data)

        logger.info(f"Obtenidas {len(data_rows)} filas de datos de '{sheet_name}'")
        return data_rows

    def update_cells(
        self,
        sheet_name: str,
        updates: List[Tuple[int, int, Any]],
        save: bool = True
    ) -> None:
        """
        Actualiza celdas específicas.

        Args:
            sheet_name: Nombre de la hoja
            updates: Lista de tuplas (row, col, value) donde row y col son 1-based
            save: Si True, guarda el archivo después de actualizar
        """
        sheet = self.get_sheet(sheet_name)

        for row, col, value in updates:
            sheet.cell(row=row, column=col, value=value)

        logger.info(f"Actualizadas {len(updates)} celdas en '{sheet_name}'")

        if save:
            self.save()

    def update_row(
        self,
        sheet_name: str,
        row_number: int,
        data: Dict[str, Any],
        save: bool = True
    ) -> None:
        """
        Actualiza una fila completa usando nombres de columnas.

        Args:
            sheet_name: Nombre de la hoja
            row_number: Número de fila (1-based)
            data: Diccionario {nombre_columna: valor}
            save: Si True, guarda el archivo después de actualizar
        """
        metadata = self.get_column_metadata(sheet_name)
        sheet = self.get_sheet(sheet_name)

        updates_count = 0
        for col_meta in metadata:
            if col_meta.name in data:
                sheet.cell(row=row_number, column=col_meta.index + 1, value=data[col_meta.name])
                updates_count += 1

        logger.info(f"Actualizada fila {row_number} en '{sheet_name}' ({updates_count} columnas)")

        if save:
            self.save()

    def append_row(
        self,
        sheet_name: str,
        data: Dict[str, Any],
        save: bool = True
    ) -> int:
        """
        Agrega una nueva fila al final de la hoja.

        Args:
            sheet_name: Nombre de la hoja
            data: Diccionario {nombre_columna: valor}
            save: Si True, guarda el archivo después de agregar

        Returns:
            Número de fila agregada (1-based)
        """
        sheet = self.get_sheet(sheet_name)
        metadata = self.get_column_metadata(sheet_name)

        # Encontrar la siguiente fila vacía
        next_row = sheet.max_row + 1

        # Escribir datos
        for col_meta in metadata:
            if col_meta.name in data:
                sheet.cell(row=next_row, column=col_meta.index + 1, value=data[col_meta.name])

        logger.info(f"Agregada fila {next_row} en '{sheet_name}'")

        if save:
            self.save()

        return next_row

    def get_column_values(
        self,
        sheet_name: str,
        column_name: str,
        start_row: int = 2,
        end_row: Optional[int] = None
    ) -> List[Any]:
        """
        Obtiene todos los valores de una columna específica.

        Args:
            sheet_name: Nombre de la hoja
            column_name: Nombre de la columna
            start_row: Fila inicial (1-based)
            end_row: Fila final (1-based, None = hasta el final)

        Returns:
            Lista de valores
        """
        metadata = self.get_column_metadata(sheet_name)
        col_meta = next((m for m in metadata if m.name == column_name), None)

        if not col_meta:
            raise ValueError(f"Columna '{column_name}' no encontrada en '{sheet_name}'")

        sheet = self.get_sheet(sheet_name)
        max_row = sheet.max_row if end_row is None else end_row

        values = []
        for row_idx in range(start_row, max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_meta.index + 1)
            if cell.value is not None:
                values.append(cell.value)

        return values

    def __enter__(self):
        """Context manager entry"""
        self.load()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.close()
        return False

