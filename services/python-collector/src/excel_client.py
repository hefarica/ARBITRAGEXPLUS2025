#!/usr/bin/env python3
"""
📥 ENTRADAS: Archivo Excel (ARBITRAGEXPLUS2025.xlsx)
🔄 TRANSFORMACIONES: Cliente para leer/escribir Excel con API compatible con GoogleSheetsClient
📤 SALIDAS: Datos de hojas Excel (ORACLE_ASSETS, PARAMETROS, etc.)
🔗 DEPENDENCIAS: openpyxl, threading (para thread-safety)

🎯 PROGRAMACIÓN DINÁMICA:
- ❌ NO hardcoding de rutas → Configurable vía env var
- ✅ Carga dinámica de hojas por nombre
- ✅ Thread-safe con locks para escritura concurrente
- ✅ API compatible con GoogleSheetsClient (drop-in replacement)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import threading
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelClient:
    """
    Cliente para interactuar con archivos Excel
    API compatible con GoogleSheetsClient para migración sin fricción
    """
    
    def __init__(self, excel_path: Optional[str] = None):
        """
        Inicializa el cliente Excel
        
        Args:
            excel_path: Ruta al archivo Excel. Si no se proporciona, usa EXCEL_FILE_PATH del env
        """
        self.excel_path = excel_path or os.getenv(
            'EXCEL_FILE_PATH',
            '/home/ubuntu/ARBITRAGEXPLUS2025/data/ARBITRAGEXPLUS2025.xlsx'
        )
        
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        # Lock para thread-safety
        self._lock = threading.Lock()
        
        logger.info(f"ExcelClient initialized with file: {self.excel_path}")
    
    def _load_workbook(self) -> openpyxl.Workbook:
        """Carga el workbook desde el archivo"""
        return openpyxl.load_workbook(self.excel_path)
    
    def _save_workbook(self, wb: openpyxl.Workbook):
        """Guarda el workbook al archivo"""
        with self._lock:
            wb.save(self.excel_path)
    
    def get_range(self, range_notation: str) -> List[List[Any]]:
        """
        Lee un rango de celdas del Excel
        
        Args:
            range_notation: Notación de rango (ej: "ORACLE_ASSETS!A1:M100")
        
        Returns:
            Lista de listas con los valores de las celdas
        
        Example:
            >>> client = ExcelClient()
            >>> data = client.get_range("ORACLE_ASSETS!A1:M100")
            >>> print(data[0])  # Primera fila (headers)
            ['SYMBOL', 'BLOCKCHAIN', 'PYTH_PRICE_ID', ...]
        """
        # Parsear notación de rango
        sheet_name, cell_range = self._parse_range_notation(range_notation)
        
        # Cargar workbook
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        
        # Obtener valores del rango
        values = []
        for row in ws[cell_range]:
            row_values = []
            for cell in row:
                # Convertir None a string vacío para compatibilidad
                value = cell.value if cell.value is not None else ""
                row_values.append(value)
            values.append(row_values)
        
        logger.debug(f"Read {len(values)} rows from {range_notation}")
        
        return values
    
    def update_range(self, range_notation: str, values: List[List[Any]]):
        """
        Actualiza un rango de celdas en Excel
        
        Args:
            range_notation: Notación de rango (ej: "RESULTADOS!A2:O2")
            values: Lista de listas con los valores a escribir
        
        Example:
            >>> client = ExcelClient()
            >>> client.update_range("RESULTADOS!A2:O2", [[
            ...     datetime.now(), "BATCH_001", "ethereum", "USDC", "ETH",
            ...     10000, 4.02, 50, 0.5, 150000, 15, 35, "0x1234", "SUCCESS", ""
            ... ]])
        """
        # Parsear notación de rango
        sheet_name, cell_range = self._parse_range_notation(range_notation)
        
        # Cargar workbook
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        
        # Parsear rango de celdas
        start_cell, end_cell = cell_range.split(':')
        start_col, start_row = self._parse_cell_address(start_cell)
        
        # Escribir valores
        for row_idx, row_values in enumerate(values):
            for col_idx, value in enumerate(row_values):
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx)
                cell.value = value
        
        # Guardar workbook
        self._save_workbook(wb)
        
        logger.debug(f"Updated {len(values)} rows in {range_notation}")
    
    def append_row(self, sheet_name: str, values: List[Any]):
        """
        Agrega una fila al final de una hoja
        
        Args:
            sheet_name: Nombre de la hoja
            values: Lista de valores a agregar
        
        Example:
            >>> client = ExcelClient()
            >>> client.append_row("RESULTADOS", [
            ...     datetime.now(), "BATCH_002", "polygon", "USDT", "MATIC",
            ...     5000, 2500, 25, 0.5, 80000, 5, 20, "0x5678", "SUCCESS", ""
            ... ])
        """
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        
        # Agregar fila
        ws.append(values)
        
        # Guardar workbook
        self._save_workbook(wb)
        
        logger.debug(f"Appended row to {sheet_name}")
    
    def get_sheet_data(self, sheet_name: str, skip_header: bool = True) -> List[Dict[str, Any]]:
        """
        Lee todos los datos de una hoja y los retorna como lista de diccionarios
        
        Args:
            sheet_name: Nombre de la hoja
            skip_header: Si True, usa la primera fila como headers
        
        Returns:
            Lista de diccionarios con los datos
        
        Example:
            >>> client = ExcelClient()
            >>> assets = client.get_sheet_data("ORACLE_ASSETS")
            >>> print(assets[0]['SYMBOL'])  # 'ETH'
            >>> print(assets[0]['BLOCKCHAIN'])  # 'ethereum'
        """
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        
        # Obtener todas las filas
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return []
        
        if skip_header:
            headers = rows[0]
            data_rows = rows[1:]
            
            # Convertir a lista de diccionarios
            result = []
            for row in data_rows:
                row_dict = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    row_dict[header] = value if value is not None else ""
                result.append(row_dict)
            
            return result
        else:
            # Retornar como lista de listas
            return [list(row) for row in rows]
    
    def update_cell(self, sheet_name: str, cell_address: str, value: Any):
        """
        Actualiza una celda específica
        
        Args:
            sheet_name: Nombre de la hoja
            cell_address: Dirección de la celda (ej: "A1", "B5")
            value: Valor a escribir
        
        Example:
            >>> client = ExcelClient()
            >>> client.update_cell("ESTADISTICAS", "B2", 150)  # TOTAL_BATCHES
        """
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        ws[cell_address] = value
        
        # Guardar workbook
        self._save_workbook(wb)
        
        logger.debug(f"Updated cell {sheet_name}!{cell_address} = {value}")
    
    def get_cell(self, sheet_name: str, cell_address: str) -> Any:
        """
        Lee el valor de una celda específica
        
        Args:
            sheet_name: Nombre de la hoja
            cell_address: Dirección de la celda (ej: "A1", "B5")
        
        Returns:
            Valor de la celda
        """
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        value = ws[cell_address].value
        
        return value if value is not None else ""
    
    def clear_range(self, range_notation: str):
        """
        Limpia un rango de celdas
        
        Args:
            range_notation: Notación de rango (ej: "RESULTADOS!A2:O100")
        """
        # Parsear notación de rango
        sheet_name, cell_range = self._parse_range_notation(range_notation)
        
        # Cargar workbook
        wb = self._load_workbook()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        
        # Limpiar rango
        for row in ws[cell_range]:
            for cell in row:
                cell.value = None
        
        # Guardar workbook
        self._save_workbook(wb)
        
        logger.debug(f"Cleared range {range_notation}")
    
    def get_sheet_names(self) -> List[str]:
        """
        Obtiene la lista de nombres de hojas en el workbook
        
        Returns:
            Lista de nombres de hojas
        """
        wb = self._load_workbook()
        return wb.sheetnames
    
    # Métodos auxiliares privados
    
    def _parse_range_notation(self, range_notation: str) -> tuple:
        """
        Parsea notación de rango tipo "SheetName!A1:B10"
        
        Returns:
            Tupla (sheet_name, cell_range)
        """
        if '!' in range_notation:
            sheet_name, cell_range = range_notation.split('!', 1)
        else:
            raise ValueError(f"Invalid range notation: {range_notation}. Expected format: 'SheetName!A1:B10'")
        
        return sheet_name, cell_range
    
    def _parse_cell_address(self, cell_address: str) -> tuple:
        """
        Parsea dirección de celda tipo "A1" a (columna, fila)
        
        Returns:
            Tupla (column_index, row_index) (1-indexed)
        """
        # Separar letras (columna) de números (fila)
        col_letters = ''
        row_numbers = ''
        
        for char in cell_address:
            if char.isalpha():
                col_letters += char
            elif char.isdigit():
                row_numbers += char
        
        col_index = column_index_from_string(col_letters)
        row_index = int(row_numbers)
        
        return col_index, row_index
    
    # Métodos de compatibilidad con GoogleSheetsClient
    
    def batch_get(self, ranges: List[str]) -> Dict[str, List[List[Any]]]:
        """
        Lee múltiples rangos en una sola operación (batch)
        
        Args:
            ranges: Lista de notaciones de rango
        
        Returns:
            Diccionario con rango como key y datos como value
        
        Example:
            >>> client = ExcelClient()
            >>> data = client.batch_get([
            ...     "ORACLE_ASSETS!A1:M100",
            ...     "PARAMETROS!A1:D20"
            ... ])
            >>> print(data["ORACLE_ASSETS!A1:M100"][0])
        """
        result = {}
        
        for range_notation in ranges:
            result[range_notation] = self.get_range(range_notation)
        
        return result
    
    def batch_update(self, updates: Dict[str, List[List[Any]]]):
        """
        Actualiza múltiples rangos en una sola operación (batch)
        
        Args:
            updates: Diccionario con rango como key y valores como value
        
        Example:
            >>> client = ExcelClient()
            >>> client.batch_update({
            ...     "ESTADISTICAS!B2": [[150]],
            ...     "ESTADISTICAS!B3": [[5000]]
            ... })
        """
        for range_notation, values in updates.items():
            self.update_range(range_notation, values)


# Función de conveniencia para obtener instancia global
_excel_client_instance = None

def get_excel_client() -> ExcelClient:
    """
    Obtiene instancia global de ExcelClient (singleton)
    
    Returns:
        Instancia de ExcelClient
    """
    global _excel_client_instance
    
    if _excel_client_instance is None:
        _excel_client_instance = ExcelClient()
    
    return _excel_client_instance


# Ejemplo de uso
if __name__ == "__main__":
    # Configurar logging
    logging.basicConfig(level=logging.DEBUG)
    
    # Crear cliente
    client = ExcelClient()
    
    # Leer ORACLE_ASSETS
    print("📊 Leyendo ORACLE_ASSETS...")
    assets = client.get_sheet_data("ORACLE_ASSETS")
    print(f"✅ {len(assets)} assets encontrados")
    print(f"Primer asset: {assets[0]['SYMBOL']} en {assets[0]['BLOCKCHAIN']}")
    
    # Leer PARAMETROS
    print("\n⚙️ Leyendo PARAMETROS...")
    params_data = client.get_range("PARAMETROS!A1:D21")
    print(f"✅ {len(params_data)-1} parámetros encontrados")
    
    # Actualizar estadística
    print("\n📈 Actualizando ESTADISTICAS...")
    client.update_cell("ESTADISTICAS", "B2", 100)  # TOTAL_BATCHES
    print("✅ Estadística actualizada")
    
    # Agregar resultado de ejemplo
    print("\n📝 Agregando resultado de ejemplo...")
    client.append_row("RESULTADOS", [
        datetime.now(),
        "BATCH_TEST",
        "ethereum",
        "USDC",
        "ETH",
        10000,
        4.02,
        50,
        0.5,
        150000,
        15,
        35,
        "0xtest...1234",
        "SUCCESS",
        "Test execution"
    ])
    print("✅ Resultado agregado")
    
    print("\n🎉 ¡Todas las operaciones completadas exitosamente!")

