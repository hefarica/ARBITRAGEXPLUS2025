"""
ExcelClient con detección automática de columnas PUSH/PULL por color de encabezado

Color azul (#4472C4) = PUSH (sistema escribe)
Color blanco (#FFFFFF) = PULL (usuario escribe, sistema lee)
"""

import openpyxl
from openpyxl.utils import column_index_from_string
from typing import List, Dict, Any, Optional, Tuple
import threading
import os

# Colores para identificación
BLUE_COLOR = "4472C4"  # PUSH
WHITE_COLOR = "FFFFFF"  # PULL

class ColumnMetadata:
    """Metadata de una columna"""
    def __init__(self, index: int, name: str, col_type: str):
        self.index = index
        self.name = name
        self.type = col_type  # "PUSH" o "PULL"
    
    def __repr__(self):
        return f"Column({self.index}, {self.name}, {self.type})"

class ExcelClientV2:
    """Cliente Excel con detección automática de PUSH/PULL"""
    
    def __init__(self, file_path: Optional[str] = None):
        self.file_path = file_path or os.getenv("EXCEL_FILE_PATH", "/home/ubuntu/ARBITRAGEXPLUS2025/data/ARBITRAGEXPLUS2025.xlsx")
        self._lock = threading.Lock()
        self._column_metadata_cache: Dict[str, List[ColumnMetadata]] = {}
        self._snapshots: Dict[str, Dict[Tuple[int, int], Any]] = {}
    
    def _load_workbook(self):
        """Carga el workbook"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        return openpyxl.load_workbook(self.file_path)
    
    def _get_header_color(self, cell) -> str:
        """Obtiene el color de fondo de una celda de encabezado"""
        if cell.fill and cell.fill.start_color:
            color = cell.fill.start_color.rgb
            if color:
                # Remover alpha channel si existe
                if len(color) == 8:
                    color = color[2:]
                return color.upper()
        return WHITE_COLOR
    
    def get_column_metadata(self, sheet_name: str, force_reload: bool = False) -> List[ColumnMetadata]:
        """Lee los headers y determina PUSH/PULL por color"""
        if not force_reload and sheet_name in self._column_metadata_cache:
            return self._column_metadata_cache[sheet_name]
        
        wb = self._load_workbook()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        metadata = []
        
        # Leer fila 1 (headers)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(1, col_idx)
            col_name = cell.value
            if not col_name:
                continue
            
            color = self._get_header_color(cell)
            col_type = "PUSH" if color == BLUE_COLOR else "PULL"
            
            metadata.append(ColumnMetadata(col_idx, str(col_name), col_type))
        
        self._column_metadata_cache[sheet_name] = metadata
        return metadata
    
    def get_pull_columns(self, sheet_name: str) -> List[ColumnMetadata]:
        """Retorna solo las columnas PULL (blancas)"""
        metadata = self.get_column_metadata(sheet_name)
        return [col for col in metadata if col.type == "PULL"]
    
    def get_push_columns(self, sheet_name: str) -> List[ColumnMetadata]:
        """Retorna solo las columnas PUSH (azules)"""
        metadata = self.get_column_metadata(sheet_name)
        return [col for col in metadata if col.type == "PUSH"]
    
    def detect_changes_in_pull_columns(self, sheet_name: str, start_row: int = 2, end_row: int = 100) -> List[Dict[str, Any]]:
        """Detecta cambios en columnas PULL comparando con snapshot"""
        wb = self._load_workbook()
        ws = wb[sheet_name]
        
        pull_columns = self.get_pull_columns(sheet_name)
        changes = []
        
        # Inicializar snapshot si no existe
        if sheet_name not in self._snapshots:
            self._snapshots[sheet_name] = {}
        
        for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
            for col_meta in pull_columns:
                cell = ws.cell(row, col_meta.index)
                current_value = cell.value
                
                snapshot_key = (row, col_meta.index)
                old_value = self._snapshots[sheet_name].get(snapshot_key)
                
                # Detectar cambio
                if current_value != old_value:
                    changes.append({
                        'row': row,
                        'column': col_meta.index,
                        'column_name': col_meta.name,
                        'old_value': old_value,
                        'new_value': current_value
                    })
                    
                    # Actualizar snapshot
                    self._snapshots[sheet_name][snapshot_key] = current_value
        
        return changes
    
    def update_push_columns(self, sheet_name: str, row: int, data: Dict[str, Any]):
        """Actualiza columnas PUSH en una fila específica"""
        with self._lock:
            wb = self._load_workbook()
            ws = wb[sheet_name]
            
            push_columns = self.get_push_columns(sheet_name)
            push_col_map = {col.name: col.index for col in push_columns}
            
            for col_name, value in data.items():
                if col_name in push_col_map:
                    col_idx = push_col_map[col_name]
                    ws.cell(row, col_idx).value = value
                    
                    # Actualizar snapshot
                    if sheet_name not in self._snapshots:
                        self._snapshots[sheet_name] = {}
                    self._snapshots[sheet_name][(row, col_idx)] = value
            
            wb.save(self.file_path)

# Singleton
_client_instance = None

def get_excel_client_v2() -> ExcelClientV2:
    """Retorna instancia singleton"""
    global _client_instance
    if _client_instance is None:
        _client_instance = ExcelClientV2()
    return _client_instance
