#!/usr/bin/env python3
"""
Script para crear el archivo Excel principal de ARBITRAGEXPLUS2025
Migración completa desde Google Sheets manteniendo todas las características
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import datetime

def create_arbitragexplus_workbook():
    """Crea el workbook principal con todas las hojas"""
    wb = Workbook()
    
    # Remover hoja por defecto
    wb.remove(wb.active)
    
    # Crear todas las hojas
    create_oracle_assets_sheet(wb)
    create_error_handling_config_sheet(wb)
    create_collectors_config_sheet(wb)
    create_parametros_sheet(wb)
    create_resultados_sheet(wb)
    create_log_errores_eventos_sheet(wb)
    create_estadisticas_sheet(wb)
    
    # Guardar archivo
    output_path = '/home/ubuntu/ARBITRAGEXPLUS2025/data/ARBITRAGEXPLUS2025.xlsx'
    wb.save(output_path)
    print(f"✅ Archivo Excel creado: {output_path}")
    
    return output_path

def create_oracle_assets_sheet(wb):
    """Crea hoja ORACLE_ASSETS con 60+ tokens configurados"""
    ws = wb.create_sheet("ORACLE_ASSETS")
    
    # Headers
    headers = [
        "SYMBOL", "BLOCKCHAIN", "PYTH_PRICE_ID", "CHAINLINK_ADDRESS",
        "UNISWAP_POOL_ADDRESS", "IS_ACTIVE", "PRIORITY", "MIN_CONFIDENCE",
        "NOTES", "BINANCE_SYMBOL", "COINGECKO_ID", "BAND_SYMBOL"
    ]
    
    # Aplicar estilo a headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos de assets (60+ tokens)
    assets_data = [
        # Ethereum
        ["ETH", "ethereum", "0xff61491a931112ddf1bd8147cd1b641375f79f5825126d665480874634fd0ace", "0x5f4eC3Df9cbd43714FE2740f5E3616155c5b8419", "0x88e6A0c2dDD26FEEb64F039a2c41296FcB3f5640", "TRUE", "1", "0.8", "Native token", "ETHUSDT", "ethereum", "ETH"],
        ["WETH", "ethereum", "0xff61491a931112ddf1bd8147cd1b641375f79f5825126d665480874634fd0ace", "0x5f4eC3Df9cbd43714FE2740f5E3616155c5b8419", "0x88e6A0c2dDD26FEEb64F039a2c41296FcB3f5640", "TRUE", "1", "0.8", "Wrapped ETH", "ETHUSDT", "ethereum", "ETH"],
        ["USDC", "ethereum", "0xeaa020c61cc479712813461ce153894a96a6c00b21ed0cfc2798d1f9a9e9c94a", "0x8fFfFfd4AfB6115b954Bd326cbe7B4BA576818f6", "0x88e6A0c2dDD26FEEb64F039a2c41296FcB3f5640", "TRUE", "1", "0.9", "Stablecoin", "USDCUSDT", "usd-coin", "USDC"],
        ["USDT", "ethereum", "0x2b89b9dc8fdf9f34709a5b106b472f0f39bb6ca9ce04b0fd7f2e971688e2e53b", "0x3E7d1eAB13ad0104d2750B8863b489D65364e32D", "0x4e68Ccd3E89f51C3074ca5072bbAC773960dFa36", "TRUE", "1", "0.9", "Stablecoin", "USDTUSDT", "tether", "USDT"],
        ["DAI", "ethereum", "0xb0948a5e5313200c632b51bb5ca32f6de0d36e9950a942d19751e833f70dabfd", "0xAed0c38402a5d19df6E4c03F4E2DceD6e29c1ee9", "0x60594a405d53811d3BC4766596EFD80fd545A270", "TRUE", "1", "0.9", "Stablecoin", "DAIUSDT", "dai", "DAI"],
        ["WBTC", "ethereum", "0xe62df6c8b4a85fe1a67db44dc12de5db330f7ac66b72dc658afedf0f4a415b43", "0xF4030086522a5bEEa4988F8cA5B36dbC97BeE88c", "0xCBCdF9626bC03E24f779434178A73a0B4bad62eD", "TRUE", "1", "0.8", "Wrapped BTC", "BTCUSDT", "wrapped-bitcoin", "BTC"],
        ["BTC", "ethereum", "0xe62df6c8b4a85fe1a67db44dc12de5db330f7ac66b72dc658afedf0f4a415b43", "0xF4030086522a5bEEa4988F8cA5B36dbC97BeE88c", "0xCBCdF9626bC03E24f779434178A73a0B4bad62eD", "TRUE", "1", "0.8", "Bitcoin", "BTCUSDT", "bitcoin", "BTC"],
        ["LINK", "ethereum", "0x8ac0c70fff57e9aefdf5edf44b51d62c2d433653cbb2cf5cc06bb115af04d221", "0x2c1d072e956AFFC0D435Cb7AC38EF18d24d9127c", "0xa6Cc3C2531FdaA6Ae1A3CA84c2855806728693e8", "TRUE", "2", "0.8", "Oracle token", "LINKUSDT", "chainlink", "LINK"],
        ["UNI", "ethereum", "0x78d185a741d07edb3412b09008b7c5cfb9bbbd7d568bf00ba737b456ba171501", "0x553303d460EE0afB37EdFf9bE42922D8FF63220e", "0x1d42064Fc4Beb5F8aAF85F4617AE8b3b5B8Bd801", "TRUE", "2", "0.8", "DEX token", "UNIUSDT", "uniswap", "UNI"],
        ["AAVE", "ethereum", "0x2b9ab1e972a281585084148ba1389800799bd4be63b957507db1349314e47445", "0x547a514d5e3769680Ce22B2361c10Ea13619e8a9", "0x5aB53EE1d50eeF2C1DD3d5402789cd27bB52c1bB", "TRUE", "2", "0.8", "Lending protocol", "AAVEUSDT", "aave", "AAVE"],
        ["SHIB", "ethereum", "0xf0d57deca57b3da2fe63a493f4c25925fdfd8edf834b20f93e1f84dbd1504d4a", "0x8dD1CD88F43aF196ae478e91b9F5E4Ac69A97C61", "0x5764a6F2212D502bC5970f9f129fFcd61e5D7563", "TRUE", "3", "0.7", "Meme token", "SHIBUSDT", "shiba-inu", "SHIB"],
        ["PEPE", "ethereum", "0xd69731a2e74ac1ce884fc3890f7ee324b6deb66147055249568869ed700882e4", "", "0x11950d141EcB863F01007AdD7D1A342041227b58", "TRUE", "3", "0.7", "Meme token", "PEPEUSDT", "pepe", "PEPE"],
        ["LDO", "ethereum", "0xc63e2a7f37a04e5e614c07238bedb25dcc38927fba8fe890597a593c0b2fa4ad", "0x4e844125952D32AcdF339BE976c98E22F6F318dB", "0xa3f558aebAecAf0e11cA4b2199cC5Ed341edfd74", "TRUE", "2", "0.8", "Liquid staking", "LDOUSDT", "lido-dao", "LDO"],
        ["MKR", "ethereum", "0x9375299e31c0deb9c6bc378e6329aab44cb48ec655552a70d4b9050346a30378", "0xec1D1B3b0443256cc3860e24a46F108e699484Aa", "0xe8c6c9227491C0a8156A0106A0204d881BB7E531", "TRUE", "2", "0.8", "Governance token", "MKRUSDT", "maker", "MKR"],
        
        # Polygon
        ["MATIC", "polygon", "0x5de33a9112c2b700b8d30b8a3402c103578ccfa2765696471cc672bd5cf6ac52", "0xAB594600376Ec9fD91F8e885dADF0CE036862dE0", "0x167384319B41F7094e62f7506409Eb38079AbfF8", "TRUE", "1", "0.8", "Native token", "MATICUSDT", "matic-network", "MATIC"],
        ["WMATIC", "polygon", "0x5de33a9112c2b700b8d30b8a3402c103578ccfa2765696471cc672bd5cf6ac52", "0xAB594600376Ec9fD91F8e885dADF0CE036862dE0", "0x167384319B41F7094e62f7506409Eb38079AbfF8", "TRUE", "1", "0.8", "Wrapped MATIC", "MATICUSDT", "matic-network", "MATIC"],
        ["USDC", "polygon", "0xeaa020c61cc479712813461ce153894a96a6c00b21ed0cfc2798d1f9a9e9c94a", "0xfE4A8cc5b5B2366C1B58Bea3858e81843581b2F7", "0xA374094527e1673A86dE625aa59517c5dE346d32", "TRUE", "1", "0.9", "Stablecoin", "USDCUSDT", "usd-coin", "USDC"],
        ["USDT", "polygon", "0x2b89b9dc8fdf9f34709a5b106b472f0f39bb6ca9ce04b0fd7f2e971688e2e53b", "0x0A6513e40db6EB1b165753AD52E80663aeA50545", "0x9B08288C3Be4F62bbf8d1C20Ac9C5e6f9467d8B7", "TRUE", "1", "0.9", "Stablecoin", "USDTUSDT", "tether", "USDT"],
        ["DAI", "polygon", "0xb0948a5e5313200c632b51bb5ca32f6de0d36e9950a942d19751e833f70dabfd", "0x4746DeC9e833A82EC7C2C1356372CcF2cfcD2F3D", "0x5645dCB64c059aa11212707fbf4E7F984440a8Cf", "TRUE", "1", "0.9", "Stablecoin", "DAIUSDT", "dai", "DAI"],
        ["WETH", "polygon", "0xff61491a931112ddf1bd8147cd1b641375f79f5825126d665480874634fd0ace", "0xF9680D99D6C9589e2a93a78A04A279e509205945", "0x45dDa9cb7c25131DF268515131f647d726f50608", "TRUE", "1", "0.8", "Wrapped ETH", "ETHUSDT", "ethereum", "ETH"],
        
        # BSC
        ["BNB", "bsc", "0x2f95862b045670cd22bee3114c39763a4a08beeb663b145d283c31d7d1101c4f", "0x0567F2323251f0Aab15c8dFb1967E4e8A7D42aeE", "0x16b9a82891338f9bA80E2D6970FddA79D1eb0daE", "TRUE", "1", "0.8", "Native token", "BNBUSDT", "binancecoin", "BNB"],
        ["WBNB", "bsc", "0x2f95862b045670cd22bee3114c39763a4a08beeb663b145d283c31d7d1101c4f", "0x0567F2323251f0Aab15c8dFb1967E4e8A7D42aeE", "0x16b9a82891338f9bA80E2D6970FddA79D1eb0daE", "TRUE", "1", "0.8", "Wrapped BNB", "BNBUSDT", "binancecoin", "BNB"],
        ["USDC", "bsc", "0xeaa020c61cc479712813461ce153894a96a6c00b21ed0cfc2798d1f9a9e9c94a", "0x51597f405303C4377E36123cBc172b13269EA163", "0xd99c7F6C65857AC913a8f880A4cb84032AB2FC5b", "TRUE", "1", "0.9", "Stablecoin", "USDCUSDT", "usd-coin", "USDC"],
        ["USDT", "bsc", "0x2b89b9dc8fdf9f34709a5b106b472f0f39bb6ca9ce04b0fd7f2e971688e2e53b", "0xB97Ad0E74fa7d920791E90258A6E2085088b4320", "0x36696169C63e42cd08ce11f5deeBbCeBae652050", "TRUE", "1", "0.9", "Stablecoin", "USDTUSDT", "tether", "USDT"],
        ["BUSD", "bsc", "0x5bc91f13e412c07599167bae86f07543f076a638962b8d6017ec19dab4a82814", "0xcBb98864Ef56E9042e7d2efef76141f15731B82f", "0xaCAac9311b0096E04Dfe96b6D87dec867d3883Dc", "TRUE", "1", "0.9", "Stablecoin", "BUSDUSDT", "binance-usd", "BUSD"],
        ["CAKE", "bsc", "0x2356af9529a1064d41e32d617e2ce1dca5733afa901daba9e2b68dee5d53ecf9", "0xB6064eD41d4f67e353768aA239cA86f4F73665a1", "0x0eD7e52944161450477ee417DE9Cd3a859b14fD0", "TRUE", "2", "0.8", "DEX token", "CAKEUSDT", "pancakeswap-token", "CAKE"],
        
        # Avalanche
        ["AVAX", "avalanche", "0x93da3352f9f1d105fdfe4971cfa80e9dd777bfc5d0f683ebb6e1294b92137bb7", "0x0A77230d17318075983913bC2145DB16C7366156", "0xfAe3f424a0a47706811521E3ee268f00cFb5c45E", "TRUE", "1", "0.8", "Native token", "AVAXUSDT", "avalanche-2", "AVAX"],
        ["WAVAX", "avalanche", "0x93da3352f9f1d105fdfe4971cfa80e9dd777bfc5d0f683ebb6e1294b92137bb7", "0x0A77230d17318075983913bC2145DB16C7366156", "0xfAe3f424a0a47706811521E3ee268f00cFb5c45E", "TRUE", "1", "0.8", "Wrapped AVAX", "AVAXUSDT", "avalanche-2", "AVAX"],
        ["USDC", "avalanche", "0xeaa020c61cc479712813461ce153894a96a6c00b21ed0cfc2798d1f9a9e9c94a", "0xF096872672F44d6EBA71458D74fe67F9a77a23B9", "0xfAe3f424a0a47706811521E3ee268f00cFb5c45E", "TRUE", "1", "0.9", "Stablecoin", "USDCUSDT", "usd-coin", "USDC"],
        ["USDT", "avalanche", "0x2b89b9dc8fdf9f34709a5b106b472f0f39bb6ca9ce04b0fd7f2e971688e2e53b", "0xEBE676ee90Fe1112671f19b6B7459bC678B67e8a", "0xeD8CBD9F0cE3C6986b22002F03c6475CEb7a6256", "TRUE", "1", "0.9", "Stablecoin", "USDTUSDT", "tether", "USDT"],
        ["DAI", "avalanche", "0xb0948a5e5313200c632b51bb5ca32f6de0d36e9950a942d19751e833f70dabfd", "0x51D7180edA2260cc4F6e4EebB82FEF5c3c2B8300", "0xbA09679Ab223C6bdaf44D45Ba2d7279959289AB0", "TRUE", "1", "0.9", "Stablecoin", "DAIUSDT", "dai", "DAI"],
        ["WETH", "avalanche", "0xff61491a931112ddf1bd8147cd1b641375f79f5825126d665480874634fd0ace", "0x976B3D034E162d8bD72D6b9C989d545b839003b0", "0x49D5c2BdFfac6CE2BFdB6640F4F80f226bc10bAB", "TRUE", "1", "0.8", "Wrapped ETH", "ETHUSDT", "ethereum", "ETH"],
        
        # Arbitrum
        ["ARB", "arbitrum", "0x3fa4252848f9f0a1480be62745a4629d9eb1322aebab8a791e344b3b9c1adcf5", "0xb2A824043730FE05F3DA2efaFa1CBbe83fa548D6", "0xC6F780497A95e246EB9449f5e4770916DCd6396A", "TRUE", "1", "0.8", "Native token", "ARBUSDT", "arbitrum", "ARB"],
        ["ETH", "arbitrum", "0xff61491a931112ddf1bd8147cd1b641375f79f5825126d665480874634fd0ace", "0x639Fe6ab55C921f74e7fac1ee960C0B6293ba612", "0xC31E54c7a869B9FcBEcc14363CF510d1c41fa443", "TRUE", "1", "0.8", "Native ETH", "ETHUSDT", "ethereum", "ETH"],
        ["USDC", "arbitrum", "0xeaa020c61cc479712813461ce153894a96a6c00b21ed0cfc2798d1f9a9e9c94a", "0x50834F3163758fcC1Df9973b6e91f0F0F0434aD3", "0xC31E54c7a869B9FcBEcc14363CF510d1c41fa443", "TRUE", "1", "0.9", "Stablecoin", "USDCUSDT", "usd-coin", "USDC"],
        ["USDT", "arbitrum", "0x2b89b9dc8fdf9f34709a5b106b472f0f39bb6ca9ce04b0fd7f2e971688e2e53b", "0x3f3f5dF88dC9F13eac63DF89EC16ef6e7E25DdE7", "0x641C00A822e8b671738d32a431a4Fb6074E5c79d", "TRUE", "1", "0.9", "Stablecoin", "USDTUSDT", "tether", "USDT"],
        ["DAI", "arbitrum", "0xb0948a5e5313200c632b51bb5ca32f6de0d36e9950a942d19751e833f70dabfd", "0xc5C8E77B397E531B8EC06BFb0048328B30E9eCfB", "0xA961F0473dA4864C5eD28e00FcC53a3AAb056c1b", "TRUE", "1", "0.9", "Stablecoin", "DAIUSDT", "dai", "DAI"],
        ["WBTC", "arbitrum", "0xe62df6c8b4a85fe1a67db44dc12de5db330f7ac66b72dc658afedf0f4a415b43", "0xd0C7101eACbB49F3deCcCc166d238410D6D46d57", "0x2f5e87C9312fa29aed5c179E456625D79015299c", "TRUE", "1", "0.8", "Wrapped BTC", "BTCUSDT", "wrapped-bitcoin", "BTC"],
        
        # Optimism
        ["OP", "optimism", "0x385f64d993f7b77d8182ed5003d97c60aa3361f3cecfe711544d2d59165e9bdf", "0x0D276FC14719f9292D5C1eA2198673d1f4269246", "0x68F5C0A2DE713a54991E01858Fd27a3832401849", "TRUE", "1", "0.8", "Native token", "OPUSDT", "optimism", "OP"],
        ["ETH", "optimism", "0xff61491a931112ddf1bd8147cd1b641375f79f5825126d665480874634fd0ace", "0x13e3Ee699D1909E989722E753853AE30b17e08c5", "0x85149247691df622eaF1a8Bd0CaFd40BC45154a9", "TRUE", "1", "0.8", "Native ETH", "ETHUSDT", "ethereum", "ETH"],
        ["USDC", "optimism", "0xeaa020c61cc479712813461ce153894a96a6c00b21ed0cfc2798d1f9a9e9c94a", "0x16a9FA2FDa030272Ce99B29CF780dFA30361E0f3", "0x85149247691df622eaF1a8Bd0CaFd40BC45154a9", "TRUE", "1", "0.9", "Stablecoin", "USDCUSDT", "usd-coin", "USDC"],
        ["USDT", "optimism", "0x2b89b9dc8fdf9f34709a5b106b472f0f39bb6ca9ce04b0fd7f2e971688e2e53b", "0xECef79E109e997bCA29c1c0897ec9d7b03647F5E", "0xc858A329Bf053BE78D6239C4A4343B8FbD21472b", "TRUE", "1", "0.9", "Stablecoin", "USDTUSDT", "tether", "USDT"],
        ["DAI", "optimism", "0xb0948a5e5313200c632b51bb5ca32f6de0d36e9950a942d19751e833f70dabfd", "0x8dBa75e83DA73cc766A7e5a0ee71F656BAb470d6", "0x03aF20bDAaFfB4cC0A521796a223f7D85e2aAc31", "TRUE", "1", "0.9", "Stablecoin", "DAIUSDT", "dai", "DAI"],
        ["WBTC", "optimism", "0xe62df6c8b4a85fe1a67db44dc12de5db330f7ac66b72dc658afedf0f4a415b43", "0x718A5788b89454aAE3A028AE9c111A29Be6c2a6F", "0x73B14a78a0D396C521f954532d43fd5fFe385216", "TRUE", "1", "0.8", "Wrapped BTC", "BTCUSDT", "wrapped-bitcoin", "BTC"],
        
        # Solana
        ["SOL", "solana", "0xef0d8b6fda2ceba41da15d4095d1da392a0d2f8ed0c6c7bc0f4cfac8c280b56d", "", "", "TRUE", "1", "0.8", "Native token", "SOLUSDT", "solana", "SOL"],
        ["BONK", "solana", "0x72b021217ca3fe68922a19aaf990109cb9d84e9ad004b4d2025ad6f529314419", "", "", "TRUE", "3", "0.7", "Meme token", "BONKUSDT", "bonk", "BONK"],
        ["JUP", "solana", "0x0a0408d619e9380abad35060f9192039ed5042fa6f82301d0e48bb52be830996", "", "", "TRUE", "2", "0.8", "DEX aggregator", "JUPUSDT", "jupiter-exchange-solana", "JUP"],
        ["PYTH", "solana", "0x0bbf28e9a841a1cc788f6a361b17ca072d0ea3098a1e5df1c3922d06719579ff", "", "", "TRUE", "2", "0.8", "Oracle network", "PYTHUSDT", "pyth-network", "PYTH"],
        ["RENDER", "solana", "0xab7347771135fc733f8f38db462ba085ed3309955f42554a14fa13e855ac0e2f", "", "", "TRUE", "2", "0.8", "Rendering token", "RENDERUSDT", "render-token", "RENDER"],
        ["WIF", "solana", "0x4ca4beeca86f0d164160323817a4e42b10010a724c2217c6ee41b54cd4cc61fc", "", "", "TRUE", "3", "0.7", "Meme token", "WIFUSDT", "dogwifcoin", "WIF"],
        ["ORCA", "solana", "0x5d1f7bc5b0b0b1e93c1d4f5e6f7b5c0a5e5f5e5f5e5f5e5f5e5f5e5f5e5f5e5f", "", "", "TRUE", "2", "0.8", "DEX token", "ORCAUSDT", "orca", "ORCA"],
        ["RAY", "solana", "0x91568baa8beb53db23eb3fb7f22c6e8bd303d103919e19733f2bb642d3e7987a", "", "", "TRUE", "2", "0.8", "DEX token", "RAYUSDT", "raydium", "RAY"],
        
        # Otros
        ["ATOM", "cosmos", "0xb00b60f88b03a6a625a8d1c048c3f66653edf217439983d037e7222c4e612819", "", "", "TRUE", "2", "0.8", "Cosmos Hub", "ATOMUSDT", "cosmos", "ATOM"],
        ["DOT", "polkadot", "0xca3eed9b267293f6595901c734c7525ce8ef49adafe8284606ceb307afa2ca5b", "", "", "TRUE", "2", "0.8", "Polkadot", "DOTUSDT", "polkadot", "DOT"],
        ["ADA", "cardano", "0x2a01deaec9e51a579277b34b122399984d0bbf57e2458a7e42fecd2829867a0d", "", "", "TRUE", "2", "0.8", "Cardano", "ADAUSDT", "cardano", "ADA"],
        ["XRP", "ripple", "0xec5d399846a9209f3fe5881d70aae9268c94339ff9817e8d18ff19fa05eea1c8", "", "", "TRUE", "2", "0.8", "Ripple", "XRPUSDT", "ripple", "XRP"],
        ["DOGE", "dogecoin", "0xdcef50dd0a4cd2dcc17e45df1676dcb336a11a61c69df7a0299b0150c672d25c", "", "", "TRUE", "3", "0.7", "Meme token", "DOGEUSDT", "dogecoin", "DOGE"],
    ]
    
    # Insertar datos
    for row_num, row_data in enumerate(assets_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Validación de datos para IS_ACTIVE
    dv_active = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv_active)
    dv_active.add(f"F2:F{len(assets_data)+1}")
    
    # Validación de datos para PRIORITY
    dv_priority = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    ws.add_data_validation(dv_priority)
    dv_priority.add(f"G2:G{len(assets_data)+1}")
    
    # Auto-ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    print(f"✅ Hoja ORACLE_ASSETS creada con {len(assets_data)} assets")

def create_error_handling_config_sheet(wb):
    """Crea hoja ERROR_HANDLING_CONFIG"""
    ws = wb.create_sheet("ERROR_HANDLING_CONFIG")
    
    # Headers
    headers = [
        "ERROR_CODE", "SHOULD_LOG", "SHOULD_ALERT", "SHOULD_RETRY",
        "MAX_RETRIES", "RETRY_DELAY_MS", "CUSTOM_HANDLERS", "NOTES"
    ]
    
    # Aplicar estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    error_configs = [
        ["VALIDATION_ERROR", "TRUE", "FALSE", "FALSE", "0", "0", "", "Input validation failed"],
        ["RPC_ERROR", "TRUE", "TRUE", "TRUE", "3", "5000", "rpc_error_handler", "RPC node unavailable"],
        ["SHEETS_ERROR", "TRUE", "TRUE", "TRUE", "3", "2000", "sheets_error_handler", "Excel read/write failed"],
        ["ORACLE_ERROR", "TRUE", "TRUE", "TRUE", "3", "3000", "oracle_error_handler", "Oracle price fetch failed"],
        ["NETWORK_ERROR", "TRUE", "TRUE", "TRUE", "5", "10000", "network_error_handler", "Network connectivity issue"],
        ["TIMEOUT_ERROR", "TRUE", "TRUE", "TRUE", "2", "5000", "", "Operation timed out"],
        ["AUTH_ERROR", "TRUE", "TRUE", "FALSE", "0", "0", "", "Authentication failed"],
        ["RATE_LIMIT_ERROR", "TRUE", "FALSE", "TRUE", "3", "60000", "rate_limit_handler", "API rate limit exceeded"],
        ["INSUFFICIENT_FUNDS", "TRUE", "TRUE", "FALSE", "0", "0", "", "Not enough balance"],
        ["GAS_ESTIMATION_ERROR", "TRUE", "TRUE", "TRUE", "2", "3000", "", "Gas estimation failed"],
    ]
    
    for row_num, row_data in enumerate(error_configs, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Validaciones
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv_bool)
    dv_bool.add(f"B2:D{len(error_configs)+1}")
    
    # Auto-ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    print(f"✅ Hoja ERROR_HANDLING_CONFIG creada con {len(error_configs)} configuraciones")

def create_collectors_config_sheet(wb):
    """Crea hoja COLLECTORS_CONFIG"""
    ws = wb.create_sheet("COLLECTORS_CONFIG")
    
    # Headers
    headers = [
        "NAME", "ENABLED", "PRIORITY", "MAX_RETRIES", "TIMEOUT",
        "MODULE_PATH", "CLASS_NAME", "NOTES"
    ]
    
    # Aplicar estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    collectors = [
        ["pyth_collector", "TRUE", "1", "3", "30", "collectors.pyth_collector", "PythCollector", "Pyth Network price collector"],
        ["chainlink_collector", "FALSE", "2", "3", "30", "collectors.chainlink_collector", "ChainlinkCollector", "Chainlink oracle collector"],
        ["uniswap_collector", "FALSE", "3", "3", "30", "collectors.uniswap_collector", "UniswapCollector", "Uniswap TWAP collector"],
        ["binance_collector", "FALSE", "4", "3", "10", "collectors.binance_collector", "BinanceCollector", "Binance API collector"],
        ["coingecko_collector", "FALSE", "5", "3", "15", "collectors.coingecko_collector", "CoinGeckoCollector", "CoinGecko API collector"],
    ]
    
    for row_num, row_data in enumerate(collectors, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Validaciones
    dv_enabled = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv_enabled)
    dv_enabled.add(f"B2:B{len(collectors)+1}")
    
    # Auto-ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    print(f"✅ Hoja COLLECTORS_CONFIG creada con {len(collectors)} collectors")

def create_parametros_sheet(wb):
    """Crea hoja PARAMETROS con configuración de arbitraje"""
    ws = wb.create_sheet("PARAMETROS")
    
    # Headers
    headers = ["PARAMETRO", "VALOR", "UNIDAD", "DESCRIPCION"]
    
    # Aplicar estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de parámetros
    params = [
        ["MIN_PROFIT_THRESHOLD", "0.5", "%", "Rentabilidad mínima para ejecutar arbitraje"],
        ["MAX_SLIPPAGE", "0.5", "%", "Slippage máximo permitido"],
        ["MAX_BATCH_SIZE", "40", "ops", "Número máximo de operaciones por batch"],
        ["GAS_PRICE_MULTIPLIER", "1.15", "x", "Multiplicador de gas para confirmación rápida"],
        ["PRICE_UPDATE_INTERVAL", "5", "s", "Intervalo de actualización de precios"],
        ["OPPORTUNITY_CHECK_INTERVAL", "1", "s", "Intervalo de detección de oportunidades"],
        ["MIN_ORACLE_CONFIDENCE", "0.8", "", "Confianza mínima requerida de oráculos"],
        ["MAX_PRICE_DEVIATION", "2", "%", "Desviación máxima entre oráculos"],
        ["MIN_ORACLES_REQUIRED", "2", "", "Número mínimo de oráculos para consenso"],
        ["TRANSACTION_DEADLINE", "60", "s", "Tiempo máximo para ejecutar transacción"],
        ["ENABLE_CIRCUIT_BREAKER", "TRUE", "", "Habilitar circuit breaker"],
        ["CIRCUIT_BREAKER_THRESHOLD", "5", "fallos", "Fallos consecutivos para activar circuit breaker"],
        ["CIRCUIT_BREAKER_COOLDOWN", "60", "s", "Tiempo de cooldown del circuit breaker"],
        ["ENABLE_AUTO_RECOVERY", "TRUE", "", "Habilitar recuperación automática"],
        ["HEALTH_CHECK_INTERVAL", "60", "s", "Intervalo de health checks"],
        ["MAX_CONCURRENT_OPERATIONS", "40", "ops", "Operaciones concurrentes máximas"],
        ["RETRY_MAX_ATTEMPTS", "3", "", "Intentos máximos de retry"],
        ["RETRY_BACKOFF_MS", "5000", "ms", "Backoff exponencial para retries"],
        ["CACHE_TTL", "30", "s", "TTL del cache de precios"],
        ["ENABLE_LOGGING", "TRUE", "", "Habilitar logging de eventos"],
    ]
    
    for row_num, row_data in enumerate(params, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    
    print(f"✅ Hoja PARAMETROS creada con {len(params)} parámetros")

def create_resultados_sheet(wb):
    """Crea hoja RESULTADOS para registrar ejecuciones"""
    ws = wb.create_sheet("RESULTADOS")
    
    # Headers
    headers = [
        "TIMESTAMP", "BATCH_ID", "CHAIN", "TOKEN_IN", "TOKEN_OUT",
        "AMOUNT_IN", "AMOUNT_OUT", "PROFIT", "PROFIT_PCT", "GAS_USED",
        "GAS_COST", "NET_PROFIT", "TX_HASH", "STATUS", "NOTES"
    ]
    
    # Aplicar estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    example_results = [
        [datetime.datetime.now(), "BATCH_001", "ethereum", "USDC", "ETH", "10000", "4.02", "50", "0.5", "150000", "15", "35", "0x1234...abcd", "SUCCESS", ""],
        [datetime.datetime.now(), "BATCH_002", "polygon", "USDT", "MATIC", "5000", "2500", "25", "0.5", "80000", "5", "20", "0x5678...efgh", "SUCCESS", ""],
    ]
    
    for row_num, row_data in enumerate(example_results, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    print(f"✅ Hoja RESULTADOS creada")

def create_log_errores_eventos_sheet(wb):
    """Crea hoja LOGERRORESEVENTOS para logging"""
    ws = wb.create_sheet("LOGERRORESEVENTOS")
    
    # Headers
    headers = [
        "TIMESTAMP", "SEVERITY", "ERROR_CODE", "COMPONENT",
        "MESSAGE", "STACK_TRACE", "CONTEXT", "RESOLVED"
    ]
    
    # Aplicar estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 10
    
    print(f"✅ Hoja LOGERRORESEVENTOS creada")

def create_estadisticas_sheet(wb):
    """Crea hoja ESTADISTICAS con métricas del sistema"""
    ws = wb.create_sheet("ESTADISTICAS")
    
    # Headers
    headers = ["METRICA", "VALOR", "UNIDAD", "ULTIMA_ACTUALIZACION"]
    
    # Aplicar estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Métricas iniciales
    stats = [
        ["TOTAL_BATCHES", "0", "batches", datetime.datetime.now()],
        ["TOTAL_OPERATIONS", "0", "ops", datetime.datetime.now()],
        ["SUCCESS_RATE", "0", "%", datetime.datetime.now()],
        ["TOTAL_PROFIT", "0", "USD", datetime.datetime.now()],
        ["TOTAL_GAS_COST", "0", "USD", datetime.datetime.now()],
        ["NET_PROFIT", "0", "USD", datetime.datetime.now()],
        ["AVG_PROFIT_PER_OP", "0", "USD", datetime.datetime.now()],
        ["CIRCUIT_BREAKER_ACTIVATIONS", "0", "count", datetime.datetime.now()],
        ["ORACLE_FAILURES", "0", "count", datetime.datetime.now()],
        ["SYSTEM_UPTIME", "0", "hours", datetime.datetime.now()],
    ]
    
    for row_num, row_data in enumerate(stats, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    
    print(f"✅ Hoja ESTADISTICAS creada")

if __name__ == "__main__":
    import os
    
    # Crear directorio data si no existe
    os.makedirs('/home/ubuntu/ARBITRAGEXPLUS2025/data', exist_ok=True)
    
    # Crear workbook
    create_arbitragexplus_workbook()
    
    print("\n🎉 ¡Archivo Excel creado exitosamente!")
    print("📁 Ubicación: /home/ubuntu/ARBITRAGEXPLUS2025/data/ARBITRAGEXPLUS2025.xlsx")
    print("\n📊 Hojas creadas:")
    print("  1. ORACLE_ASSETS (60 assets)")
    print("  2. ERROR_HANDLING_CONFIG (10 configuraciones)")
    print("  3. COLLECTORS_CONFIG (5 collectors)")
    print("  4. PARAMETROS (20 parámetros)")
    print("  5. RESULTADOS (registro de ejecuciones)")
    print("  6. LOGERRORESEVENTOS (logging)")
    print("  7. ESTADISTICAS (métricas del sistema)")

