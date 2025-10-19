"""
Script para crear la hoja BLOCKCHAINS con columnas coloreadas:
- Columna B (NAME): BLANCA (PULL) - Usuario escribe nombre de blockchain
- Columnas A y C-AY: AZULES (PUSH) - Sistema escribe automáticamente
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

# Ruta del Excel
EXCEL_PATH = "/home/ubuntu/ARBITRAGEXPLUS2025/data/ARBITRAGEXPLUS2025.xlsx"

# Definición de columnas para BLOCKCHAINS
BLOCKCHAIN_COLUMNS = [
    ("BLOCKCHAIN_ID", "PUSH"),      # A
    ("NAME", "PULL"),                # B - Usuario escribe aquí
    ("CHAIN_ID", "PUSH"),            # C
    ("NATIVE_TOKEN", "PUSH"),        # D
    ("SYMBOL", "PUSH"),              # E
    ("RPC_URL_1", "PUSH"),           # F
    ("RPC_URL_2", "PUSH"),           # G
    ("RPC_URL_3", "PUSH"),           # H
    ("WSS_URL", "PUSH"),             # I
    ("EXPLORER_URL", "PUSH"),        # J
    ("BLOCK_TIME_MS", "PUSH"),       # K
    ("GAS_PRICE_GWEI", "PUSH"),      # L
    ("MAX_GAS_PRICE", "PUSH"),       # M
    ("MIN_GAS_PRICE", "PUSH"),       # N
    ("EIP1559_SUPPORTED", "PUSH"),   # O
    ("BASE_FEE", "PUSH"),            # P
    ("PRIORITY_FEE", "PUSH"),        # Q
    ("GAS_LIMIT", "PUSH"),           # R
    ("MULTICALL_ADDRESS", "PUSH"),   # S
    ("WETH_ADDRESS", "PUSH"),        # T
    ("USDC_ADDRESS", "PUSH"),        # U
    ("USDT_ADDRESS", "PUSH"),        # V
    ("DAI_ADDRESS", "PUSH"),         # W
    ("SUPPORTED_DEXES", "PUSH"),     # X
    ("SUPPORTED_PROTOCOLS", "PUSH"), # Y
    ("TVL_USD", "PUSH"),             # Z
    ("DAILY_VOLUME_USD", "PUSH"),    # AA
    ("TRANSACTION_COUNT", "PUSH"),   # AB
    ("AVERAGE_GAS_COST", "PUSH"),    # AC
    ("FINALITY_BLOCKS", "PUSH"),     # AD
    ("REORG_PROTECTION", "PUSH"),    # AE
    ("MEV_PROTECTION", "PUSH"),      # AF
    ("FLASHBOTS_SUPPORTED", "PUSH"), # AG
    ("PRIVATE_TX_SUPPORTED", "PUSH"),# AH
    ("HEALTH_STATUS", "PUSH"),       # AI
    ("LAST_BLOCK_NUMBER", "PUSH"),   # AJ
    ("LAST_BLOCK_TIMESTAMP", "PUSH"),# AK
    ("SYNC_STATUS", "PUSH"),         # AL
    ("LATENCY_MS", "PUSH"),          # AM
    ("SUCCESS_RATE", "PUSH"),        # AN
    ("ERROR_RATE", "PUSH"),          # AO
    ("RETRY_COUNT", "PUSH"),         # AP
    ("TIMEOUT_MS", "PUSH"),          # AQ
    ("MAX_RETRIES", "PUSH"),         # AR
    ("CIRCUIT_BREAKER_THRESHOLD", "PUSH"), # AS
    ("RATE_LIMIT_PER_SECOND", "PUSH"),     # AT
    ("NOTES", "PUSH"),               # AU
    ("CREATED_AT", "PUSH"),          # AV
    ("UPDATED_AT", "PUSH"),          # AW
    ("IS_ACTIVE", "PUSH"),           # AX
    ("AVG_GAS_USED", "PUSH"),        # AY
]

def create_blockchains_sheet():
    """Crea o actualiza la hoja BLOCKCHAINS con columnas coloreadas"""
    
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: No se encontró {EXCEL_PATH}")
        return
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Eliminar hoja BLOCKCHAINS si existe
    if "BLOCKCHAINS" in wb.sheetnames:
        del wb["BLOCKCHAINS"]
        print("🗑️  Hoja BLOCKCHAINS anterior eliminada")
    
    # Crear nueva hoja
    ws = wb.create_sheet("BLOCKCHAINS", 0)  # Insertar al inicio
    
    # Colores
    BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    # Escribir headers con colores
    for col_idx, (col_name, col_type) in enumerate(BLOCKCHAIN_COLUMNS, 1):
        cell = ws.cell(1, col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT if col_type == "PUSH" else Font(bold=True, size=11)
        cell.fill = BLUE_FILL if col_type == "PUSH" else WHITE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columna
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    
    # Agregar filas de ejemplo
    example_rows = [
        ["ethereum", ""],  # Fila 2: Solo NAME (columna B)
        ["polygon", ""],   # Fila 3: Solo NAME (columna B)
        ["bsc", ""],       # Fila 4: Solo NAME (columna B)
    ]
    
    for row_idx, row_data in enumerate(example_rows, 2):
        # Solo escribir en columna B (NAME)
        ws.cell(row_idx, 2).value = row_data[0]
    
    # Proteger hoja (opcional - permite edición solo en columna B)
    # ws.protection.sheet = True
    # ws.protection.password = "arbitrage2025"
    
    wb.save(EXCEL_PATH)
    print(f"✅ Hoja BLOCKCHAINS creada con {len(BLOCKCHAIN_COLUMNS)} columnas")
    print(f"   - Columna B (NAME): BLANCA (PULL) - Usuario escribe aquí")
    print(f"   - Columnas A, C-AY: AZULES (PUSH) - Sistema escribe automáticamente")
    print(f"   - 3 filas de ejemplo agregadas")

if __name__ == "__main__":
    create_blockchains_sheet()
