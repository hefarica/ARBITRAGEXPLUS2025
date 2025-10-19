#!/usr/bin/env python3
"""
Script para clonar EXACTAMENTE el Google Sheets a Excel
Replica todas las hojas, columnas y datos
"""

import sys
import os
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configuración
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SERVICE_ACCOUNT_FILE = '/home/ubuntu/ARBITRAGEXPLUS2025/keys/gsheets-sa.json'
SPREADSHEET_ID = '1qLKS8anyP8lb9jCVujT6KzTPjaSjNrAPYWhCxv4sChQ'
OUTPUT_FILE = '/home/ubuntu/ARBITRAGEXPLUS2025/data/ARBITRAGEXPLUS2025_FULL.xlsx'

def main():
    print("🔄 Clonando Google Sheets a Excel...")
    print(f"📥 Origen: Google Sheets ID {SPREADSHEET_ID}")
    print(f"📤 Destino: {OUTPUT_FILE}\n")
    
    # Conectar a Google Sheets
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    
    # Obtener metadata
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets = spreadsheet['sheets']
    
    print(f"📊 Total de hojas encontradas: {len(sheets)}\n")
    
    # Crear workbook de Excel
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Procesar cada hoja
    for idx, sheet_info in enumerate(sheets, 1):
        sheet_title = sheet_info['properties']['title']
        print(f"[{idx}/{len(sheets)}] Procesando: {sheet_title}")
        
        try:
            # Leer datos de la hoja
            result = service.spreadsheets().values().get(
                spreadsheetId=SPREADSHEET_ID,
                range=f"{sheet_title}!A1:ZZZ10000"  # Rango amplio
            ).execute()
            values = result.get('values', [])
            
            if not values:
                print(f"   ⚠️  Hoja vacía, creando solo headers")
                values = [[]]
            
            # Crear hoja en Excel
            ws = wb.create_sheet(title=sheet_title)
            
            # Escribir datos
            for row_idx, row in enumerate(values, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Formatear header (primera fila)
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"   ✅ {len(values)} filas x {len(values[0]) if values else 0} columnas")
            
        except Exception as e:
            print(f"   ❌ Error: {e}")
            continue
    
    # Guardar Excel
    print(f"\n💾 Guardando archivo Excel...")
    wb.save(OUTPUT_FILE)
    
    # Estadísticas finales
    file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)  # MB
    print(f"\n✅ ¡Clonación completada!")
    print(f"📁 Archivo: {OUTPUT_FILE}")
    print(f"📏 Tamaño: {file_size:.2f} MB")
    print(f"📊 Hojas: {len(wb.sheetnames)}")
    print(f"📋 Hojas creadas: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()

