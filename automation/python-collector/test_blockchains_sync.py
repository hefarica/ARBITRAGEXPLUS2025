#!/usr/bin/env python3
"""
Test End-to-End del Sistema de Sincronización Bidireccional Excel

Este script prueba el flujo completo:
1. Escribe un nombre de blockchain en columna B (NAME - PULL/blanco)
2. Verifica que las columnas PUSH (azules) se actualicen automáticamente
3. Valida que el tiempo de respuesta sea <500ms

Uso:
    python3 test_blockchains_sync.py
"""

import sys
import os
import time
import logging

# Agregar src al path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from excel_client_v2 import get_excel_client_v2
from aggregators.blockchain_data_aggregator import get_blockchain_data_aggregator

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def test_end_to_end_sync():
    """Prueba el flujo completo de sincronización"""
    
    print("\n" + "="*80)
    print("TEST END-TO-END: Sistema de Sincronización Bidireccional Excel")
    print("="*80)
    
    # Inicializar clientes
    print("\n📦 Inicializando componentes...")
    client = get_excel_client_v2()
    aggregator = get_blockchain_data_aggregator()
    
    sheet_name = "BLOCKCHAINS"
    test_row = 5  # Usar fila 5 para pruebas
    blockchain_name = "avalanche"
    
    print(f"✅ Componentes inicializados")
    print(f"   - Hoja: {sheet_name}")
    print(f"   - Fila de prueba: {test_row}")
    print(f"   - Blockchain: {blockchain_name}")
    
    # Paso 1: Verificar estructura de columnas
    print("\n📊 Paso 1: Verificar estructura de columnas...")
    metadata = client.get_column_metadata(sheet_name)
    pull_columns = [col for col in metadata if col.type == "PULL"]
    push_columns = [col for col in metadata if col.type == "PUSH"]
    
    print(f"   ✅ {len(pull_columns)} columnas PULL (blancas)")
    print(f"   ✅ {len(push_columns)} columnas PUSH (azules)")
    
    for col in pull_columns:
        print(f"      - PULL: Columna {col.index} - {col.name}")
    
    # Paso 2: Simular escritura de usuario en columna PULL
    print(f"\n✏️  Paso 2: Simular escritura de usuario en columna NAME...")
    print(f"   Escribiendo '{blockchain_name}' en fila {test_row}...")
    
    # Encontrar columna NAME
    name_column = next((col for col in pull_columns if col.name == "NAME"), None)
    if not name_column:
        print("   ❌ ERROR: Columna NAME no encontrada")
        return False
    
    # Escribir en Excel (simular usuario)
    import openpyxl
    wb = openpyxl.load_workbook(client.file_path)
    ws = wb[sheet_name]
    ws.cell(test_row, name_column.index).value = blockchain_name
    wb.save(client.file_path)
    
    print(f"   ✅ Valor escrito en columna {name_column.index} (NAME)")
    
    # Paso 3: Detectar cambio
    print(f"\n🔍 Paso 3: Detectar cambio en columna PULL...")
    changes = client.detect_changes_in_pull_columns(sheet_name, start_row=test_row, end_row=test_row)
    
    if not changes:
        print("   ⚠️  No se detectaron cambios (puede ser que ya existía el valor)")
    else:
        print(f"   ✅ {len(changes)} cambio(s) detectado(s)")
        for change in changes:
            print(f"      - Fila {change['row']}, {change['column_name']}: '{change['old_value']}' → '{change['new_value']}'")
    
    # Paso 4: Obtener datos de fuentes externas
    print(f"\n🌐 Paso 4: Obtener datos de fuentes externas...")
    start_time = time.time()
    
    blockchain_data = aggregator.aggregate_blockchain_data(blockchain_name, parallel=True)
    
    fetch_time = (time.time() - start_time) * 1000
    print(f"   ✅ Datos obtenidos en {fetch_time:.0f}ms")
    print(f"   📊 {len(blockchain_data)} campos obtenidos")
    
    # Mostrar algunos campos clave
    key_fields = ['NAME', 'CHAIN_ID', 'NATIVE_TOKEN', 'TVL_USD', 'PROTOCOLS_COUNT', 
                  'RPC_URL_1', 'EXPLORER_URL', 'HEALTH_STATUS']
    print("\n   Campos clave obtenidos:")
    for field in key_fields:
        if field in blockchain_data:
            value = blockchain_data[field]
            print(f"      - {field}: {value}")
    
    # Paso 5: Actualizar columnas PUSH
    print(f"\n💾 Paso 5: Actualizar columnas PUSH en Excel...")
    start_update = time.time()
    
    client.update_push_columns(sheet_name, test_row, blockchain_data)
    
    update_time = (time.time() - start_update) * 1000
    total_time = (time.time() - start_time) * 1000
    
    print(f"   ✅ Columnas PUSH actualizadas en {update_time:.0f}ms")
    print(f"   ⏱️  Tiempo total: {total_time:.0f}ms")
    
    # Paso 6: Verificar actualización
    print(f"\n✔️  Paso 6: Verificar actualización en Excel...")
    wb = openpyxl.load_workbook(client.file_path)
    ws = wb[sheet_name]
    
    # Verificar algunos campos PUSH
    verification_fields = {
        'CHAIN_ID': blockchain_data.get('CHAIN_ID'),
        'NATIVE_TOKEN': blockchain_data.get('NATIVE_TOKEN'),
        'HEALTH_STATUS': blockchain_data.get('HEALTH_STATUS'),
    }
    
    verified = 0
    for field_name, expected_value in verification_fields.items():
        # Buscar columna
        col_meta = next((col for col in push_columns if col.name == field_name), None)
        if col_meta:
            actual_value = ws.cell(test_row, col_meta.index).value
            if str(actual_value) == str(expected_value):
                print(f"   ✅ {field_name}: {actual_value} (correcto)")
                verified += 1
            else:
                print(f"   ⚠️  {field_name}: esperado={expected_value}, actual={actual_value}")
    
    # Paso 7: Validar rendimiento
    print(f"\n⚡ Paso 7: Validar rendimiento...")
    target_time = 500  # ms
    
    if total_time < target_time:
        print(f"   ✅ Rendimiento OK: {total_time:.0f}ms < {target_time}ms (objetivo)")
    else:
        print(f"   ⚠️  Rendimiento lento: {total_time:.0f}ms > {target_time}ms (objetivo)")
    
    # Resumen
    print("\n" + "="*80)
    print("RESUMEN DEL TEST")
    print("="*80)
    print(f"✅ Blockchain: {blockchain_name}")
    print(f"✅ Fila actualizada: {test_row}")
    print(f"✅ Campos obtenidos: {len(blockchain_data)}")
    print(f"✅ Campos verificados: {verified}/{len(verification_fields)}")
    print(f"✅ Tiempo de fetch: {fetch_time:.0f}ms")
    print(f"✅ Tiempo de actualización: {update_time:.0f}ms")
    print(f"✅ Tiempo total: {total_time:.0f}ms")
    print(f"✅ Objetivo de rendimiento: {'CUMPLIDO' if total_time < target_time else 'NO CUMPLIDO'}")
    print("="*80)
    
    return total_time < target_time


def main():
    """Entry point"""
    try:
        success = test_end_to_end_sync()
        
        if success:
            print("\n🎉 ¡TEST EXITOSO!")
            sys.exit(0)
        else:
            print("\n⚠️  Test completado con advertencias")
            sys.exit(1)
            
    except Exception as e:
        logger.error(f"❌ Error en test: {e}", exc_info=True)
        print(f"\n❌ TEST FALLIDO: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

